import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pypdf import PdfReader
import streamlit as st

st.set_page_config(page_title="Stücklisten Extraktor", page_icon="📊")

st.title("📊 Stromlaufplan Stücklisten-Extraktor")
st.write("Laden Sie Ihr Stromlaufplan-PDF hoch, um automatisch eine formatierte Excel-Stückliste zu generieren.")

uploaded_file = st.file_uploader("PDF-Datei auswählen", type=["pdf"])

def extract_tables_and_text(pdf_file):
    reader = PdfReader(pdf_file)
    rows = []
    
    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue
        
        lines = text.split('\n')
        for line in lines:
            line_str = line.strip()
            if not line_str:
                continue
            parts = re.split(r'\s{2,}|\t', line_str)
            if len(parts) >= 2:
                rows.append(parts)
    return rows

def create_excel(rows):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stückliste"
    ws.views.sheetView[0].showGridLines = True
    
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    data_font = Font(name='Segoe UI', size=10)
    zebra_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    border_side = Side(style='thin', color='D9D9D9')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = data_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output)
    output.seek(0)
    return output

if uploaded_file is not None:
    with st.spinner("PDF wird verarbeitet..."):
        try:
            rows = extract_tables_and_text(uploaded_file)
            if not rows:
                st.warning("Keine tabellarischen Daten im PDF gefunden.")
            else:
                excel_data = create_excel(rows)
                st.success("Stückliste erfolgreich extrahiert!")
                
                st.download_button(
                    label="📥 Excel-Datei herunterladen",
                    data=excel_data,
                    file_name="Stueckliste.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Fehler bei der Verarbeitung: {e}")
